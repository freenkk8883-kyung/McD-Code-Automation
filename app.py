import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import chardet
import io
import os
from datetime import datetime

# ─────────────────────────────────────────────
# 페이지 설정
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="RAW Data Automator",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─────────────────────────────────────────────
# 스타일
# ─────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@300;400;500;600&family=Noto+Sans+KR:wght@300;400;500;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Noto Sans KR', sans-serif;
}

/* 전체 배경 */
.stApp { background-color: #0d0f14; color: #e2e8f0; }

/* 사이드바 */
[data-testid="stSidebar"] {
    background-color: #13161e;
    border-right: 1px solid #252a38;
}

/* 헤더 */
.app-header {
    display: flex;
    align-items: center;
    gap: 12px;
    padding: 4px 0 24px 0;
    border-bottom: 1px solid #252a38;
    margin-bottom: 28px;
}
.app-badge {
    background: #4f9cf9;
    color: #0a0f1e;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 11px;
    font-weight: 700;
    padding: 4px 12px;
    letter-spacing: 0.1em;
}
.app-title { font-size: 20px; font-weight: 700; color: #e2e8f0; margin: 0; }
.app-sub { font-family: 'IBM Plex Mono', monospace; font-size: 11px; color: #505869; margin: 0; }

/* 매체 카드 그리드 */
.media-grid {
    display: grid;
    grid-template-columns: repeat(2, 1fr);
    gap: 10px;
    margin-bottom: 24px;
}
.media-card {
    background: #13161e;
    border: 1px solid #252a38;
    padding: 14px 16px;
    position: relative;
    transition: border-color 0.2s;
}
.media-card.active {
    border-color: #4ade80;
    background: rgba(74,222,128,0.04);
}
.media-card.active::before {
    content: '';
    position: absolute;
    top: 0; left: 0;
    width: 3px; height: 100%;
    background: #4ade80;
}
.media-card-name {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 11px;
    font-weight: 600;
    color: #8892a4;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-bottom: 4px;
}
.media-card.active .media-card-name { color: #4ade80; }
.media-card-status {
    font-size: 12px;
    color: #505869;
}
.media-card.active .media-card-status { color: #e2e8f0; }
.media-card-time {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 10px;
    color: #505869;
    margin-top: 4px;
}

/* 알림 토스트 */
.toast {
    background: #13161e;
    border: 1px solid #4ade80;
    border-left: 4px solid #4ade80;
    padding: 14px 18px;
    margin-bottom: 12px;
    display: flex;
    align-items: center;
    gap: 12px;
    animation: slideIn 0.3s ease;
}
@keyframes slideIn {
    from { opacity: 0; transform: translateX(-10px); }
    to   { opacity: 1; transform: translateX(0); }
}
.toast-icon { font-size: 18px; }
.toast-text { flex: 1; }
.toast-title { font-size: 13px; font-weight: 600; color: #4ade80; }
.toast-sub { font-size: 11px; color: #8892a4; font-family: 'IBM Plex Mono', monospace; }

/* 섹션 레이블 */
.section-label {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 10px;
    text-transform: uppercase;
    letter-spacing: 0.12em;
    color: #505869;
    margin-bottom: 8px;
}

/* 파일업로더 커스텀 */
[data-testid="stFileUploader"] {
    background: #13161e;
    border: 1px dashed #252a38;
    padding: 8px;
}

/* 버튼 */
.stButton > button {
    background: #4f9cf9;
    color: #0a0f1e;
    border: none;
    font-family: 'Noto Sans KR', sans-serif;
    font-weight: 700;
    font-size: 13px;
    padding: 10px 24px;
    width: 100%;
    border-radius: 0;
}
.stButton > button:hover { background: #3d7dd4; }

/* 데이터프레임 */
[data-testid="stDataFrame"] { background: #13161e; }

/* 구분선 */
hr { border-color: #252a38; }

/* 메트릭 */
[data-testid="metric-container"] {
    background: #13161e;
    border: 1px solid #252a38;
    padding: 12px 16px;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# 코드 사전 로드 (session_state 캐싱)
# ─────────────────────────────────────────────

# ─────────────────────────────────────────────
# 내장 코드 사전 (코드 사전 파일 없이도 동작)
# ─────────────────────────────────────────────
BUILTIN_MEDIA_MAP = {
    'ADPA0':'AD Packer','APPI0':'Appier','APPL0':'Apple','BLIN0':'Blind',
    'CASH0':'Cashwalk','COUP0':'CoupangPlay','DAAN0':'Daangn','EVER0':'Everytime',
    'GOOG0':'Google UAC','GOOG1':'Google UACe','HIGH0':'HIghClass',
    'KAKA0':'Kakao Mobility','KAKA1':'Kakao Moment','KAKA2':'Kakao DMP',
    'KAKA3':'Kakao Navi','Kakao':'Kakao Moment','KAKAO':'Kakao Moment',
    'Kakaotalk':'Kakao Moment','Kakao_DMP':'Kakao DMP',
    'KIDS0':'Kidsnote','LGTV0':'LG TV','MOLO0':'Moloco',
    'MOTI0':'Motiv_CTV','Motive':'Motiv_CTV','Motiv':'Motiv_CTV','Motive CTV':'Motiv_CTV',
    'NAVE0':'Naver GFA','GFA':'Naver GFA','GFA_DMP':'Naver GFA_DMP',
    'NAVE1':'Naver Map','NAVE2':'Naver NOSP','NAVE3':'Naver Search',
    'NAVE4':'Naver SA','NAVE5':'Naver GFA_DMP','Naver GFA DMP':'Naver GFA_DMP',
    'NETF0':'Netflix','NASM0':'Nasmedia CTV','MILI0':'Milianze',
    'OKCA0':'OKCashbag','OPGG0':'OPGG','PODB0':'Podbbang','REME0':'Remember',
    'RUND0':'Runday','SAMS0':'Samsung CTV','SAMS1':'Samsung TV','SKPL0':'SK Planet',
    'SMR0':'SMR','SPOT0':'SPOTV','TEAD0':'Teads','TIKT0':'Tiktok','TOSS0':'Toss',
    'TTD0':'TTD','TTDC0':'TTD CTV','TVIN0':'Tving',
    'TWIT0':'Twitter','TWIT1':'Twitter','X':'Twitter','X(Twitter)':'Twitter',
    'YOUT0':'Youtube','CHZZ0':'CHZZK','치지직':'CHZZK','카카오 네비':'Kakao Navi',
    'kakaonavi':'Kakao Navi',
}

BUILTIN_PRODUCT_MAP = {
    'WORL0':'WORLDCUP2026','5SSK0':'5s skip','ADDR0':'ADDRESSABLE',
    'APPI0':'Appier Retargeting','APPI1':'Appier UA','APPL0':'Apple search',
    'AUDI0':'Audio Pre-roll','BIGB0':'Big banner',
    'BIZB0':'Bizboard','BIZB1':'Bizboard CPT','BIZB2':'Bizboard DMP',
    'BUMP0':'Bumper','BUND0':'Bundle',
    'COMM0':'Comment Banner','COMM1':'Community Banner','COMM2':'Community Native',
    'CTV0':'CTV','CTVI0':'CTV inStream','CURA0':'Curation','CUST0':'Custom Line-up',
    'DEMA0':'Demand GEN','DISP0':'Display','DOUB0':'Double Crown',
    'ENDB0':'End banner','ENDI0':'Ending Popup','EVEN0':'Event banner',
    'FAST0':'FAST','FEED0':'Feed','FEED1':'Feed Carousel','FEED2':'Feed_DMP',
    'FOCU0':'Focus Board','FOCU1':'Focus Full view','FOOD0':'Food Curation',
    'FPM0':'FPM','FPMS0':'FPM_Shorts','GOOG0':'Google UACe',
    'HOME0':'Home Banner','HOME1':'Home card','HOME2':'Home bottom banner',
    'HOME3':'Home Wrapping MO','HOME4':'Home Wrapping PC',
    'INFE0':'InFeed','INFE1':'In-feed ad',
    'INRE0':'inRead Vertical','INRE1':'InRead Vertical Display',
    'INST0':'Instream Video','INTE0':'Interstitial (front banner)',
    'KBO10':'KBO 1 Month PKG','KBOA0':'KBO All PKG','KBOP0':'KBO Postseason',
    'LGCT0':'LG CTV Homescreen Masthead','LIST0':'List','LIST1':'List Banner',
    'LIVE0':'Live Sports',
    'M.BR0':'M.Branding','M.BR1':'M.Branding DMP','M.BR2':'M.BrandingDA',
    'M.SM0':'M.Smart Channel',
    'MAIN0':'Main Banner','MAIN1':'Main enter feed 1st','MAIN2':'Main sports feed 1st',
    'MAIN3':'Main Banner','MAIN4':'Main pop-up',
    'MAST0':'Masthead','MAST1':'Masthead CPH','MLBL0':'MLB Live CM',
    'MOBI0':'Mobile App netword DA','MOBI1':'Mobile Web Big Banner',
    'MOME0':'Moment Banner','MILL0':'MilliTalk Banner',
    'NATI0':'Native','NATI1':'Native','NATI2':'Native AD',
    'NATI3':'Native AD DMP','NATI4':'Native App','NATI5':'Native Banner',
    'NAVE0':'Naver Map','NETW0':'Network DA',
    'NONT0':'Non Target CPV','NONS0':'NonSkip','NONS1':'Nonskip_DMP',
    'NORT0':'North-mid america','OQUI0':'O Quiz','O&OA0':'O&O Ads_Masthead Only',
    'PARI0':'Paris Olympics PKG','PAUS0':'Pause',
    'PHOT0':'PhotoPopup','PHOT1':'Photoswipe','POI0':'POI',
    'POPU0':'Pop-up Banner','POWE0':'Powerlink','PREM0':'Premium Photo Ending',
    'PTW0':'PTW','PWT0':'PWT','QTON0':'QTONE',
    'R&F0':'R&F','R&FT0':'R&F TopFeed','RICH0':'Richpop All Day',
    'ROAD0':'Roadblock','RUNO0':'Run of shorts','SEAR0':'Search Banner',
    'SELE0':'Select Line-up','SMAR0':'Smart Channel','SMAR1':'Smart channel Full screen',
    'SMRL0':'SMR Live','SNSV0':'SNS Virtual Advertising','SOV10':'SOV 100%',
    'SPEC0':'Special DA','SPEC1':'Special DA Expandable',
    'SPEC2':'Special DA_Video Expandable_V2',
    'SPLA0':'Splash','SPLA1':'Splash AD','SPLA2':'Splash banner',
    'SPOR0':'SportsDA','SQUA0':'Square banner','STAN0':'Standard (Belt Banner)',
    'TAKE0':'Takeover Bundle','TAKE1':'Takeover',
    'THUM0':'Thumbnail','THUM1':'Thumbnail Banner',
    'TIME0':'Timeboard','TIME1':'Timeboard_Video Expandable',
    'TODA0':"Today's TVINGTOP20",'TOP10':'Top 10','TOP20':'Top 20',
    'TOPF0':'TopFeed','TOPV0':'TopView',
    'TRV0':'TRV','TRVD0':'TRV DMP','TRVR0':'TRV Reach',
    'TTDC0':'TTD CTV (Samsung TV, LG Ads)','TTDC1':'TTD CTV (SK Btv, U+tv, KT)',
    'TVI0':'TVI','TVID0':'TVI DMP','UAC0':'UAC','UACA0':'UAC Action','UACE0':'UACe',
    'VERT0':'Vertical','VERT1':'Vertical Video Takeover','VIDE0':'Video DA',
    'VOD0':'VOD','VODP0':'VOD Pre-roll',
    'VRC0':'VRC','VRCD0':'VRC DMP','VRCT0':'VRC TF 3.0',
    'VRC(0':'VRC(Bumper)','VRC(1':'VRC(Shorts)',
    'VVC0':'VVC','VVC(0':'VVC (15s)','VVCD0':'VVC DMP',
    'VVC30':'VVC_30s(Trueview Instream)','VVCS0':'VVC_Shorts',
    'WEBT0':'Webtoon Big Banner','WEBT1':'Webtoon Bigbanner','WEBT2':'Webtoon Special DA',
    'DIRE0':'Directions End Banner','MYCA0':'My Car Care Service Button',
    'CHZZ0':'CHZZK','CHIJ0':'Chijijik','CHAL0':'Challenger','CHEC0':'Check-in Banner',
    'EXIT0':'Exit pop-up','KORE0':'Korean League CM',
    'KAKA0':'Kakaotalk Plus Friends','YEAR0':'Year-end Awards Package',
}

@st.cache_data
def load_code_dict(uploaded_dict_file):
    # 내장 코드 사전으로 시작
    media_map = dict(BUILTIN_MEDIA_MAP)
    product_map = dict(BUILTIN_PRODUCT_MAP)
    if uploaded_dict_file is None:
        return media_map, product_map
    # 업로드된 파일로 덮어쓰기 (최신 우선)
    raw = pd.read_excel(uploaded_dict_file, sheet_name='DATA RAW', header=None)
    for i in range(1, len(raw)):
        media  = str(raw.iloc[i, 2]).strip()
        final  = str(raw.iloc[i, 6]).strip()
        alias  = str(raw.iloc[i, 8]).strip()
        if media not in ('nan','') and final not in ('nan',''):
            media_map[final] = media
            if alias not in ('nan',''):
                for a in alias.split(','):
                    a = a.strip()
                    if a: media_map[a] = media
    for i in range(1, len(raw)):
        product = str(raw.iloc[i, 9]).strip()
        final   = str(raw.iloc[i, 13]).strip()
        if product not in ('nan','') and final not in ('nan',''):
            product_map[final] = product
    return media_map, product_map

# ─────────────────────────────────────────────
# 매체 자동 감지
# ─────────────────────────────────────────────
def detect_media(filename, file_bytes):
    fname = filename.lower()

    # 1순위: 파일명 패턴
    if fname == 'result.csv' or (fname.startswith('result') and fname.endswith('.csv')):
        return 'naver_gfa'
    if 'daily-' in fname and fname.endswith('.xlsx'):
        return 'twitter'
    if fname.startswith('dmc_') and '보고서' in fname:
        return 'kakao'

    # 2순위: 인코딩 + 컬럼명
    enc = chardet.detect(file_bytes[:3000])['encoding'] or 'utf-8'
    try:
        if 'utf-16' in enc.lower():
            text = file_bytes.decode('utf-16')
        else:
            text = file_bytes.decode('utf-8-sig', errors='ignore')
        lines = text.split('\n')
        # GFA는 헤더가 0번째 줄, Youtube는 2번째 줄
        header = lines[0] + (lines[2] if len(lines) > 2 else '')

        if 'TrueView 조회수' in header or '동영상 25% 재생' in header:
            return 'youtube'
        if '광고 소재 이름' in header and ('총 비용' in header or '총비용' in header):
            return 'naver_gfa'
        if '소재 이름' in header and '클릭률' in header:
            return 'kakao'
        if '트윗 참여' in header or '리트윗 수' in header:
            return 'twitter'
    except:
        pass

    # xlsx 컬럼 확인
    if fname.endswith('.xlsx'):
        try:
            df = pd.read_excel(io.BytesIO(file_bytes), header=0)
            cols = ' '.join(df.columns.tolist())
            if '트윗 참여' in cols: return 'twitter'
            if 'TrueView' in cols: return 'youtube'
            if '광고 소재 이름' in cols: return 'naver_gfa'
            if '소재 이름' in cols: return 'kakao'
        except:
            pass

    return 'unknown'

MEDIA_INFO = {
    'youtube':   {'label': 'Youtube',    'emoji': '▶', 'color': '#f87171'},
    'naver_gfa': {'label': 'Naver GFA',  'emoji': 'N', 'color': '#4ade80'},
    'kakao':     {'label': 'Kakao',      'emoji': 'K', 'color': '#fbbf24'},
    'twitter':   {'label': 'Twitter (X)','emoji': 'X', 'color': '#4f9cf9'},
    'unknown':   {'label': '알 수 없음',  'emoji': '?', 'color': '#8892a4'},
}

COST_FORMULA = {
    'youtube':   lambda v: f'={v}/0.9',
    'naver_gfa': lambda v: f'={v}/1.1*0.95',
    'kakao':     lambda v: f'={v}*0.95',
    'twitter':   lambda v: f'={v}/0.9',
}

# ─────────────────────────────────────────────
# 파서
# ─────────────────────────────────────────────
CREATIVE_KO = {
    'Snackwrap':'스낵랩','Churros':'츄러스','Churros1':'츄러스1',
    'Churros2':'츄러스2','Shake':'쉐이크','Platform':'플랫폼',
    'Carousel':'DA','Carousel_v2':'DA','Carousel_v3':'DA','DA':'DA',
}
WEEK_RANGES = [
    ('W1','2026-02-03','2026-02-11'),('W2','2026-02-12','2026-02-19'),
    ('W3','2026-02-20','2026-02-25'),('W4','2026-02-26','2026-03-04'),
    ('W5','2026-03-05','2026-03-11'),('W6','2026-03-12','2026-03-18'),
    ('W7','2026-03-19','2026-03-25'),('W8','2026-03-26','2026-04-01'),
    ('W9','2026-04-02','2026-04-08'),('W10','2026-04-09','2026-04-15'),
    ('W11','2026-04-16','2026-04-22'),('W12','2026-04-23','2026-04-29'),
    ('W13','2026-04-30','2026-05-06'),('W14','2026-05-07','2026-05-13'),
    ('W15','2026-05-14','2026-05-20'),
]

def get_week(d):
    d = str(d)[:10].replace('.','-').rstrip('-')
    for w,s,e in WEEK_RANGES:
        if s <= d <= e: return w
    return ''

def parse_date(d):
    return str(d)[:10].replace('.','-').rstrip('-')

def clean_num(v):
    v = str(v).strip().strip('"').replace(',','').replace('%','').strip()
    return '' if v in ['-','nan','--',''] else v

def parse_campaign_code(code, media_map, product_map):
    # 캠페인명 예: 2602_NAVE0_M.BR0_BA Q1
    # [0]날짜 [1]매체코드 [2]상품코드 [3~]캠페인명
    parts = str(code).split('_')
    if len(parts) < 3: return '', ''
    media_code   = parts[1]
    product_code = parts[2]
    return media_map.get(media_code,''), product_map.get(product_code,'')

def parse_creative_code(code):
    parts = str(code).split('_')
    if not parts or parts[0] == '': return '', ''

    # 형식 A: M_Snackwrap_H_15_A (텍소노미 소재 코드)
    # → [0]=M, [1]=소재, [2]=방향, [3]=초수
    if parts[0] in ('M','A','C','P') and len(parts) >= 4:
        creative = CREATIVE_KO.get(parts[1], parts[1])
        sec = parts[3] if parts[3].isdigit() else ''
        return creative, sec

    # 형식 B: VVC_스낵랩_15s_가로 (광고 이름 직접 파싱)
    # → [0]=상품, [1]=소재, [2]=초수s_방향
    if len(parts) >= 3:
        creative_raw = parts[1]
        creative = CREATIVE_KO.get(creative_raw, creative_raw)
        sec_raw = parts[2].replace('s','').split('_')[0]
        sec = sec_raw if sec_raw.isdigit() else ''
        return creative, sec

    # 형식 C: Carousel_v2, DA 처럼 소재명 자체인 경우
    creative = CREATIVE_KO.get(parts[0], '')
    if creative:
        return creative, ''

    return '', ''


def parse_product_from_ad(ad_name, product_map):
    """광고 이름에서 상품명 추출 (텍소노미 코드 없는 파일용)"""
    parts = str(ad_name).split('_')
    if not parts: return ''
    # VVC_스낵랩_... → VVC0 → product_map 조회
    product_code = parts[0] + '0'
    return product_map.get(product_code, parts[0])

RAW_COLS = ['미디어','상품명','캠페인','그룹','광고','소재','초수','일자',
            '비용(net)','노출','클릭','조회','25','50','75','100',
            '0.25','0.5','0.75','1','인게이지','리트윗','답글수','팔로우','미디어참여','투표','Week']
AUTO_COLS = {'미디어','상품명','소재','초수','Week'}

COL_MAP = {
    'youtube':   {'campaign':'캠페인','group':'광고그룹','ad':'광고 이름','date':'일','cost':'net',
                  'imp':'노출수','click':'클릭수','view':'TrueView 조회수',
                  'm25':'동영상 25% 재생','m50':'동영상 50% 재생','m75':'동영상 75% 재생','m100':'동영상 100% 재생'},
    'naver_gfa': {'campaign':'캠페인 이름','group':'광고 그룹 이름','ad':'광고 소재 이름',
                  'date':'기간','cost':'총비용','imp':'노출수','click':'클릭수'},
    'kakao':     {'campaign':'캠페인 이름','group':'광고그룹 이름','ad':'소재 이름',
                  'date':'일','cost':'비용','imp':'노출수','click':'클릭수'},
    'twitter':   {'campaign':'캠페인 이름','group':'광고 그룹 이름','ad':'광고 이름',
                  'date':'기간','cost':'비용','imp':'노출수','click':'클릭수',
                  'engage':'트윗 참여','retweet':'리트윗 수','reply':'답글',
                  'follow':'팔로우 수','media_eng':'미디어 참여'},
}

def extract_campaigns(rows):
    """업로드된 매체 데이터에서 캠페인명 추출 (GFA 필터링 기준)"""
    campaigns = set()
    for row in rows:
        c = str(row.get('캠페인', '')).strip()
        if c:
            campaigns.add(c)
    return campaigns


def parse_file(media_key, file_bytes, media_map, product_map):
    rows = []
    cm = COL_MAP.get(media_key, {})

    if media_key == 'youtube':
        enc = chardet.detect(file_bytes[:3000])['encoding'] or 'utf-16'
        text = file_bytes.decode('utf-16' if 'utf-16' in enc.lower() else enc)
        lines = text.split('\n')
        headers = lines[2].strip().split('\t')
        for line in lines[3:]:
            cols = line.strip().split('\t')
            if len(cols) < 10: continue
            row = {headers[i]: cols[i] for i in range(min(len(headers),len(cols)))}
            creative_code = str(row.get('캠페인','')).strip()
            ad_name = str(row.get('광고 이름','')).strip()
            creative, sec = parse_creative_code(creative_code)
            # 상품명: 캠페인코드에서 추출 → 실패 시 광고 이름에서 폴백
            _, product = parse_campaign_code(creative_code, media_map, product_map)
            if not product:
                product = parse_product_from_ad(ad_name, product_map)
            date_raw = parse_date(row.get('일',''))
            out = {c:'' for c in RAW_COLS}
            out.update({'미디어':'Youtube','상품명':product,'캠페인':creative_code,
                        '그룹':str(row.get('광고그룹','')).strip(),'광고':ad_name,
                        '소재':creative,'초수':sec,'일자':date_raw,
                        '비용(net)':clean_num(row.get('net','')) or clean_num(row.get('비용','')),
                        '노출':clean_num(row.get('노출수','')),'클릭':clean_num(row.get('클릭수','')),
                        '조회':clean_num(row.get('TrueView 조회수','')),
                        '25':clean_num(row.get('동영상 25% 재생','')),'50':clean_num(row.get('동영상 50% 재생','')),
                        '75':clean_num(row.get('동영상 75% 재생','')),'100':clean_num(row.get('동영상 100% 재생','')),
                        'Week':get_week(date_raw)})
            rows.append(out)

    elif media_key == 'naver_gfa':
        df = pd.read_csv(io.BytesIO(file_bytes), encoding='utf-8-sig')
        # 다른 매체(유튜브/카카오)에서 추출한 캠페인명 기준으로 필터링
        ref = st.session_state.get('ref_campaigns', set())
        if ref:
            # 기준 캠페인명과 GFA 캠페인명을 비교 (부분 매칭)
            # 예: 유튜브 캠페인 '2602_YOUT0_VRC0_BA Q1' → 'BA Q1' 추출
            # GFA 캠페인 '2602_NAVE0_M.BR0_BA Q1' → 'BA Q1' 포함 여부 확인
            ref_keywords = set()
            for c in ref:
                parts = str(c).split('_')
                # 텍소노미 형식: 마지막 부분이 캠페인명
                if len(parts) >= 4:
                    ref_keywords.add('_'.join(parts[3:]))  # 예: 'BA Q1'
                else:
                    ref_keywords.add(c)
            def match_campaign(gfa_campaign):
                for kw in ref_keywords:
                    if kw in str(gfa_campaign):
                        return True
                return False
            df = df[df[cm['campaign']].apply(match_campaign)]
        else:
            # 다른 매체 없으면 전체 데이터 사용 (필터링 없음)
            pass
        for _, row in df.iterrows():
            campaign = str(row.get(cm['campaign'],''))
            ad = str(row.get(cm['ad'],''))
            media, product = parse_campaign_code(campaign, media_map, product_map)
            creative, sec = parse_creative_code(ad)
            date_raw = parse_date(row.get(cm['date'],''))
            cost = clean_num(row.get(cm['cost'],''))
            out = {c:'' for c in RAW_COLS}
            out.update({'미디어':media or 'Naver GFA','상품명':product,'캠페인':campaign,
                        '그룹':str(row.get(cm['group'],'')),'광고':ad,
                        '소재':creative,'초수':sec,'일자':date_raw,
                        '비용(net)':cost,
                        '노출':clean_num(row.get(cm['imp'],'')),
                        '클릭':clean_num(row.get(cm['click'],'')),
                        'Week':get_week(date_raw)})
            rows.append(out)

    elif media_key == 'kakao':
        enc = chardet.detect(file_bytes[:3000])['encoding'] or 'utf-16'
        text = file_bytes.decode('utf-16' if 'utf-16' in enc.lower() else enc)
        df = pd.read_csv(io.StringIO(text), sep='\t', quotechar='"')
        for _, row in df.iterrows():
            campaign = str(row.get(cm['campaign'],''))
            ad = str(row.get(cm['ad'],''))
            media, product = parse_campaign_code(campaign, media_map, product_map)
            creative, sec = parse_creative_code(ad)
            date_raw = parse_date(row.get(cm['date'],''))
            cost = clean_num(str(row.get(cm['cost'],'')))
            out = {c:'' for c in RAW_COLS}
            out.update({'미디어':media or 'Kakao Moment','상품명':product,'캠페인':campaign,
                        '그룹':str(row.get(cm['group'],'')),'광고':ad,
                        '소재':creative,'초수':sec,'일자':date_raw,
                        '비용(net)':cost,'노출':clean_num(str(row.get(cm['imp'],''))),'클릭':clean_num(str(row.get(cm['click'],''))),
                        'Week':get_week(date_raw)})
            rows.append(out)

    elif media_key == 'twitter':
        df = pd.read_excel(io.BytesIO(file_bytes), header=0)
        for _, row in df.iterrows():
            campaign = str(row.get(cm['campaign'],''))
            ad = str(row.get(cm['ad'],''))
            date_raw = parse_date(row.get(cm['date'],''))
            cost = clean_num(str(row.get(cm['cost'],'')))
            group = str(row.get(cm['group'],''))
            product = 'Vertical Video Takeover' if 'VVT' in campaign or 'Vertical' in group else 'PTW'
            out = {c:'' for c in RAW_COLS}
            out.update({'미디어':'Twitter','상품명':product,'캠페인':campaign,
                        '그룹':group,'광고':ad,'소재':'DA','일자':date_raw,
                        '비용(net)':cost,'노출':clean_num(str(row.get(cm['imp'],''))),'클릭':clean_num(str(row.get(cm['click'],''))),
                        '인게이지':clean_num(str(row.get(cm['engage'],''))),
                        '리트윗':clean_num(str(row.get(cm['retweet'],''))),
                        '답글수':clean_num(str(row.get(cm['reply'],''))),
                        '팔로우':clean_num(str(row.get(cm['follow'],''))),
                        '미디어참여':clean_num(str(row.get(cm['media_eng'],''))),
                        'Week':get_week(date_raw)})
            rows.append(out)

    return rows

# ─────────────────────────────────────────────
# 엑셀 출력
# ─────────────────────────────────────────────
def to_excel_bytes_merged(all_data):
    """여러 매체 데이터를 하나의 엑셀로 합쳐서 출력"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Raw Data'

    hf_auto  = PatternFill('solid', start_color='FFF0C040')
    hf_input = PatternFill('solid', start_color='FF1F3D6B')
    hfa  = Font(name='Arial', bold=True, color='FF000000', size=10)
    hfi  = Font(name='Arial', bold=True, color='FFFFFFFF', size=10)
    af   = PatternFill('solid', start_color='FFFFF8E0')
    afont= Font(name='Arial', size=9, color='FF805500')
    nfont= Font(name='Arial', size=9)
    cfont= Font(name='Arial', size=9, color='FF000080')
    thin = Side(style='thin', color='FFCCCCCC')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 헤더
    for ci, cn in enumerate(RAW_COLS, 1):
        cell = ws.cell(row=1, column=ci, value=cn)
        cell.fill = hf_auto if cn in AUTO_COLS else hf_input
        cell.font = hfa if cn in AUTO_COLS else hfi
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bdr

    # 전체 매체 데이터 순서대로 합치기
    current_row = 2
    for mk, data in all_data.items():
        cost_fn = COST_FORMULA.get(mk)
        for rd in data['rows']:
            for ci, cn in enumerate(RAW_COLS, 1):
                val = rd[cn]
                if cn == '비용(net)' and cost_fn and val:
                    try:
                        cell = ws.cell(row=current_row, column=ci, value=cost_fn(float(val)))
                        cell.font = cfont
                    except:
                        cell = ws.cell(row=current_row, column=ci, value=val)
                        cell.font = nfont
                else:
                    if cn in ('노출','클릭','조회') and val:
                        try: val = int(float(val))
                        except: pass
                    cell = ws.cell(row=current_row, column=ci, value=val)
                    cell.fill = af if cn in AUTO_COLS else PatternFill()
                    cell.font = afont if cn in AUTO_COLS else nfont
                cell.border = bdr
                cell.alignment = Alignment(vertical='center')
            current_row += 1

    col_widths = {'미디어':14,'상품명':16,'캠페인':34,'그룹':24,'광고':22,'소재':10,'초수':6,
                  '일자':12,'비용(net)':18,'노출':12,'클릭':10,'조회':10,'25':8,'50':8,'75':8,
                  '100':8,'0.25':6,'0.5':6,'0.75':6,'1':6,'인게이지':10,'리트윗':8,
                  '답글수':8,'팔로우':8,'미디어참여':10,'투표':8,'Week':8}
    for ci, cn in enumerate(RAW_COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(cn, 10)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_excel_bytes(rows, media_key):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Raw Data'
    hf_auto  = PatternFill('solid', start_color='FFF0C040')
    hf_input = PatternFill('solid', start_color='FF1F3D6B')
    hfa  = Font(name='Arial', bold=True, color='FF000000', size=10)
    hfi  = Font(name='Arial', bold=True, color='FFFFFFFF', size=10)
    af   = PatternFill('solid', start_color='FFFFF8E0')
    afont= Font(name='Arial', size=9, color='FF805500')
    nfont= Font(name='Arial', size=9)
    cfont= Font(name='Arial', size=9, color='FF000080')
    thin = Side(style='thin', color='FFCCCCCC')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, cn in enumerate(RAW_COLS, 1):
        cell = ws.cell(row=1, column=ci, value=cn)
        cell.fill = hf_auto if cn in AUTO_COLS else hf_input
        cell.font = hfa if cn in AUTO_COLS else hfi
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bdr

    cost_fn = COST_FORMULA.get(media_key)
    for ri, rd in enumerate(rows, 2):
        for ci, cn in enumerate(RAW_COLS, 1):
            val = rd[cn]
            if cn == '비용(net)' and cost_fn and val:
                try:
                    cell = ws.cell(row=ri, column=ci, value=cost_fn(float(val)))
                    cell.font = cfont
                except:
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.font = nfont
            else:
                if cn in ('노출','클릭','조회') and val:
                    try: val = int(float(val))
                    except: pass
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill = af if cn in AUTO_COLS else PatternFill()
                cell.font = afont if cn in AUTO_COLS else nfont
            cell.border = bdr
            cell.alignment = Alignment(vertical='center')

    col_widths = {'미디어':14,'상품명':16,'캠페인':34,'그룹':24,'광고':22,'소재':10,'초수':6,
                  '일자':12,'비용(net)':18,'노출':12,'클릭':10,'조회':10,'25':8,'50':8,'75':8,
                  '100':8,'0.25':6,'0.5':6,'0.75':6,'1':6,'인게이지':10,'리트윗':8,
                  '답글수':8,'팔로우':8,'미디어참여':10,'투표':8,'Week':8}
    for ci, cn in enumerate(RAW_COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(cn, 10)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ─────────────────────────────────────────────
# session_state 초기화
# ─────────────────────────────────────────────
if 'uploaded_media' not in st.session_state:
    st.session_state.uploaded_media = {}   # {media_key: {rows, filename, time, row_count}}
if 'toasts' not in st.session_state:
    st.session_state.toasts = []           # 최근 업데이트 알림 목록
if 'ref_campaigns' not in st.session_state:
    st.session_state.ref_campaigns = set() # 유튜브/카카오에서 추출한 캠페인명 기준값

# ─────────────────────────────────────────────
# 사이드바
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div class="section-label">코드 사전</div>', unsafe_allow_html=True)

    dict_file = st.file_uploader(
        "코드 사전 파일 업로드 (선택)",
        type=['xlsx'],
        key='dict_uploader',
        help="맥도날드_자동화_작업용.xlsx — 업로드 시 최신 코드로 자동 업데이트"
    )
    media_map, product_map = load_code_dict(dict_file)

    if dict_file:
        st.success(f"✓ 업로드 코드 사전 사용 중\n매체 {len(media_map)}개 / 상품 {len(product_map)}개")
    else:
        st.info(f"✓ 내장 코드 사전 사용 중\n매체 {len(media_map)}개 / 상품 {len(product_map)}개")

    # ── 코드 사전 내용 확인 & 다운로드 ──
    with st.expander("📋 코드 사전 확인 / 다운로드"):
        tab_m, tab_p = st.tabs(["매체 코드", "상품 코드"])

        with tab_m:
            df_media = pd.DataFrame([
                {'최종CODE': k, '매체명': v}
                for k, v in media_map.items()
                if len(k) <= 8  # 코드 형식만 표시
            ]).sort_values('최종CODE').reset_index(drop=True)
            st.dataframe(df_media, use_container_width=True, height=200)

        with tab_p:
            df_product = pd.DataFrame([
                {'최종CODE': k, '상품명': v}
                for k, v in product_map.items()
            ]).sort_values('최종CODE').reset_index(drop=True)
            st.dataframe(df_product, use_container_width=True, height=200)

        # 현재 코드 사전을 엑셀로 다운로드
        def make_dict_excel():
            wb = openpyxl.Workbook()

            # Media 시트
            ws_m = wb.active
            ws_m.title = 'Media'
            ws_m.append(['최종CODE', '매체명'])
            for k, v in sorted(media_map.items()):
                ws_m.append([k, v])

            # Product 시트
            ws_p = wb.create_sheet('Product')
            ws_p.append(['최종CODE', '상품명'])
            for k, v in sorted(product_map.items()):
                ws_p.append([k, v])

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        st.download_button(
            label="⬇  현재 코드 사전 다운로드",
            data=make_dict_excel(),
            file_name=f"코드사전_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            key='download_dict',
            use_container_width=True
        )

    st.divider()
    st.markdown('<div class="section-label">업로드 현황</div>', unsafe_allow_html=True)

    all_media = ['youtube','naver_gfa','kakao','twitter']
    for mk in all_media:
        info = MEDIA_INFO[mk]
        uploaded = mk in st.session_state.uploaded_media
        if uploaded:
            data = st.session_state.uploaded_media[mk]
            st.markdown(f"""
            <div class="media-card active">
                <div class="media-card-name">{info['emoji']} {info['label']}</div>
                <div class="media-card-status">✓ {data['row_count']}행 업데이트</div>
                <div class="media-card-time">{data['time']}</div>
            </div>""", unsafe_allow_html=True)
        else:
            st.markdown(f"""
            <div class="media-card">
                <div class="media-card-name">{info['emoji']} {info['label']}</div>
                <div class="media-card-status">대기 중</div>
            </div>""", unsafe_allow_html=True)

    # ── GFA 필터링 기준 캠페인명 표시 ──
    if st.session_state.ref_campaigns:
        st.divider()
        st.markdown('<div class="section-label">GFA 필터링 기준</div>', unsafe_allow_html=True)
        st.caption("유튜브/카카오에서 자동 추출된 캠페인명")
        for c in sorted(st.session_state.ref_campaigns):
            parts = str(c).split('_')
            label = '_'.join(parts[3:]) if len(parts) >= 4 else c
            badge = f'<div style="background:#0d0f14;border:1px solid #252a38;padding:6px 10px;margin-bottom:4px;font-family:monospace;font-size:10px;color:#4ade80;">✓ {label}</div>'
            st.markdown(badge, unsafe_allow_html=True)
        if st.button("기준 초기화", key="clear_campaigns", use_container_width=True):
            st.session_state.ref_campaigns = set()
            st.rerun()

# ─────────────────────────────────────────────
# 메인 영역
# ─────────────────────────────────────────────
st.markdown("""
<div class="app-header">
    <div class="app-badge">McD</div>
    <div>
        <p class="app-title">RAW Data Automator</p>
        <p class="app-sub">매체 리포트 → RAW 데이터 자동 변환</p>
    </div>
</div>
""", unsafe_allow_html=True)

# 업데이트 알림 토스트
if st.session_state.toasts:
    for toast in st.session_state.toasts[-3:]:  # 최근 3개만
        info = MEDIA_INFO.get(toast['media_key'], MEDIA_INFO['unknown'])
        st.markdown(f"""
        <div class="toast">
            <div class="toast-icon">✅</div>
            <div class="toast-text">
                <div class="toast-title">{info['label']} 데이터 업데이트 완료</div>
                <div class="toast-sub">{toast['filename']} · {toast['row_count']}행 · {toast['time']}</div>
            </div>
        </div>""", unsafe_allow_html=True)

# 파일 업로드 영역
st.markdown('<div class="section-label">매체 파일 업로드</div>', unsafe_allow_html=True)
st.caption("파일명 또는 내부 컬럼을 분석하여 매체를 자동으로 감지합니다.")

uploaded_files = st.file_uploader(
    "매체 리포트 파일을 업로드하세요 (여러 파일 동시 가능)",
    type=['csv', 'xlsx', 'xls'],
    accept_multiple_files=True,
    key='media_uploader'
)

if uploaded_files:
    for uf in uploaded_files:
        file_bytes = uf.read()
        media_key = detect_media(uf.name, file_bytes)
        info = MEDIA_INFO[media_key]

        col1, col2 = st.columns([3, 1])
        with col1:
            if media_key == 'unknown':
                st.warning(f"⚠️ `{uf.name}` — 매체를 자동 감지하지 못했습니다. 아래에서 직접 선택해주세요.")
                media_key = st.selectbox(
                    "매체 선택",
                    options=list(MEDIA_INFO.keys())[:-1],
                    format_func=lambda x: MEDIA_INFO[x]['label'],
                    key=f'select_{uf.name}'
                )
            else:
                st.success(f"**{info['label']}** 파일 감지 — `{uf.name}`")

        with col2:
            if st.button("변환 실행", key=f'btn_{uf.name}'):
                with st.spinner(f"{info['label']} 데이터 처리 중..."):
                    try:
                        rows = parse_file(media_key, file_bytes, media_map, product_map)
                        now = datetime.now().strftime('%m/%d %H:%M')
                        st.session_state.uploaded_media[media_key] = {
                            'rows': rows,
                            'filename': uf.name,
                            'time': now,
                            'row_count': len(rows),
                            'media_key': media_key,
                        }
                        # 유튜브/카카오 업로드 시 캠페인명 기준값 누적 추출
                        if media_key in ('youtube', 'kakao'):
                            new_camps = extract_campaigns(rows)
                            st.session_state.ref_campaigns.update(new_camps)
                        st.session_state.toasts.append({
                            'media_key': media_key,
                            'filename': uf.name,
                            'row_count': len(rows),
                            'time': now,
                        })
                        st.rerun()
                    except Exception as e:
                        st.error(f"오류 발생: {e}")

st.divider()

# 결과 및 다운로드 영역
if st.session_state.uploaded_media:
    st.markdown('<div class="section-label">변환 결과</div>', unsafe_allow_html=True)

    # ── 전체 통합 다운로드 ──
    total_rows = sum(d['row_count'] for d in st.session_state.uploaded_media.values())
    media_labels = ' + '.join(MEDIA_INFO[mk]['label'] for mk in st.session_state.uploaded_media)
    today = datetime.now().strftime('%Y%m%d')

    st.markdown(f"""
    <div style="background:#13161e;border:1px solid #4f9cf9;padding:16px 20px;
                display:flex;align-items:center;justify-content:space-between;margin-bottom:16px;">
        <div>
            <div style="font-family:'IBM Plex Mono',monospace;font-size:10px;
                        color:#4f9cf9;text-transform:uppercase;letter-spacing:0.1em;margin-bottom:4px;">
                전체 통합 다운로드
            </div>
            <div style="font-size:13px;color:#e2e8f0;">
                {media_labels} &nbsp;·&nbsp; 총 <b>{total_rows}행</b>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    merged_bytes = to_excel_bytes_merged(st.session_state.uploaded_media)
    st.download_button(
        label=f"⬇  전체 매체 통합 RAW 데이터 다운로드 ({total_rows}행)",
        data=merged_bytes,
        file_name=f"RAW_Data_ALL_{today}.xlsx",
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        key='download_all',
        type='primary'
    )

    st.divider()

    # ── 매체별 탭 ──
    tabs = st.tabs([MEDIA_INFO[mk]['label'] for mk in st.session_state.uploaded_media.keys()])
    for tab, (mk, data) in zip(tabs, st.session_state.uploaded_media.items()):
        with tab:
            col1, col2, col3 = st.columns(3)
            col1.metric("총 행수", f"{data['row_count']}행")
            col2.metric("파일명", data['filename'][:20])
            col3.metric("업데이트", data['time'])

            df_preview = pd.DataFrame(data['rows'])
            preview_cols = ['미디어','상품명','캠페인','소재','초수','일자','Week','비용(net)','노출','클릭']
            available_cols = [c for c in preview_cols if c in df_preview.columns]
            st.dataframe(
                df_preview[available_cols].head(20),
                use_container_width=True,
                height=300
            )

            excel_bytes = to_excel_bytes(data['rows'], mk)
            today = datetime.now().strftime('%Y%m%d')
            st.download_button(
                label=f"⬇  {MEDIA_INFO[mk]['label']} RAW 데이터 다운로드",
                data=excel_bytes,
                file_name=f"RAW_Data_{MEDIA_INFO[mk]['label']}_{today}.xlsx",
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                key=f'download_{mk}'
            )
else:
    st.markdown("""
    <div style="text-align:center;padding:48px;color:#505869;border:1px dashed #252a38;">
        <div style="font-size:32px;margin-bottom:12px;">📂</div>
        <div style="font-family:'IBM Plex Mono',monospace;font-size:12px;">
            매체 파일을 업로드하면 여기에 결과가 표시됩니다
        </div>
    </div>
    """, unsafe_allow_html=True)
