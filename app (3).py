import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import chardet
import io
import os
from datetime import datetime

# ─────────────────────────────────────────────
# 페이지 설정
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="RAW Data Automator",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─────────────────────────────────────────────
# 스타일
# ─────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@300;400;500;600&family=Noto+Sans+KR:wght@300;400;500;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Noto Sans KR', sans-serif;
}

/* 전체 배경 */
.stApp { background-color: #0d0f14; color: #e2e8f0; }

/* 사이드바 */
[data-testid="stSidebar"] {
    background-color: #13161e;
    border-right: 1px solid #252a38;
}

/* 헤더 */
.app-header {
    display: flex;
    align-items: center;
    gap: 12px;
    padding: 4px 0 24px 0;
    border-bottom: 1px solid #252a38;
    margin-bottom: 28px;
}
.app-badge {
    background: #4f9cf9;
    color: #0a0f1e;
    font-family: 'IBM Plex Mono', monospace;
    font-size: 11px;
    font-weight: 700;
    padding: 4px 12px;
    letter-spacing: 0.1em;
}
.app-title { font-size: 20px; font-weight: 700; color: #e2e8f0; margin: 0; }
.app-sub { font-family: 'IBM Plex Mono', monospace; font-size: 11px; color: #505869; margin: 0; }

/* 매체 카드 그리드 */
.media-grid {
    display: grid;
    grid-template-columns: repeat(2, 1fr);
    gap: 10px;
    margin-bottom: 24px;
}
.media-card {
    background: #13161e;
    border: 1px solid #252a38;
    padding: 14px 16px;
    position: relative;
    transition: border-color 0.2s;
}
.media-card.active {
    border-color: #4ade80;
    background: rgba(74,222,128,0.04);
}
.media-card.active::before {
    content: '';
    position: absolute;
    top: 0; left: 0;
    width: 3px; height: 100%;
    background: #4ade80;
}
.media-card-name {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 11px;
    font-weight: 600;
    color: #8892a4;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-bottom: 4px;
}
.media-card.active .media-card-name { color: #4ade80; }
.media-card-status {
    font-size: 12px;
    color: #505869;
}
.media-card.active .media-card-status { color: #e2e8f0; }
.media-card-time {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 10px;
    color: #505869;
    margin-top: 4px;
}

/* 알림 토스트 */
.toast {
    background: #13161e;
    border: 1px solid #4ade80;
    border-left: 4px solid #4ade80;
    padding: 14px 18px;
    margin-bottom: 12px;
    display: flex;
    align-items: center;
    gap: 12px;
    animation: slideIn 0.3s ease;
}
@keyframes slideIn {
    from { opacity: 0; transform: translateX(-10px); }
    to   { opacity: 1; transform: translateX(0); }
}
.toast-icon { font-size: 18px; }
.toast-text { flex: 1; }
.toast-title { font-size: 13px; font-weight: 600; color: #4ade80; }
.toast-sub { font-size: 11px; color: #8892a4; font-family: 'IBM Plex Mono', monospace; }

/* 섹션 레이블 */
.section-label {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 10px;
    text-transform: uppercase;
    letter-spacing: 0.12em;
    color: #505869;
    margin-bottom: 8px;
}

/* 파일업로더 커스텀 */
[data-testid="stFileUploader"] {
    background: #13161e;
    border: 1px dashed #252a38;
    padding: 8px;
}

/* 버튼 */
.stButton > button {
    background: #4f9cf9;
    color: #0a0f1e;
    border: none;
    font-family: 'Noto Sans KR', sans-serif;
    font-weight: 700;
    font-size: 13px;
    padding: 10px 24px;
    width: 100%;
    border-radius: 0;
}
.stButton > button:hover { background: #3d7dd4; }

/* 데이터프레임 */
[data-testid="stDataFrame"] { background: #13161e; }

/* 구분선 */
hr { border-color: #252a38; }

/* 메트릭 */
[data-testid="metric-container"] {
    background: #13161e;
    border: 1px solid #252a38;
    padding: 12px 16px;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# 코드 사전 로드 (session_state 캐싱)
# ─────────────────────────────────────────────
@st.cache_data
def load_code_dict(uploaded_dict_file):
    raw = pd.read_excel(uploaded_dict_file, sheet_name='DATA RAW', header=None)
    media_map, product_map = {}, {}
    for i in range(1, len(raw)):
        media  = str(raw.iloc[i, 2]).strip()
        final  = str(raw.iloc[i, 6]).strip()
        alias  = str(raw.iloc[i, 8]).strip()
        if media not in ('nan','') and final not in ('nan',''):
            media_map[final] = media
            if alias not in ('nan',''):
                for a in alias.split(','):
                    a = a.strip()
                    if a: media_map[a] = media
    for i in range(1, len(raw)):
        product = str(raw.iloc[i, 9]).strip()
        final   = str(raw.iloc[i, 13]).strip()
        if product not in ('nan','') and final not in ('nan',''):
            product_map[final] = product
    return media_map, product_map

# ─────────────────────────────────────────────
# 매체 자동 감지
# ─────────────────────────────────────────────
def detect_media(filename, file_bytes):
    fname = filename.lower()

    # 1순위: 파일명 패턴
    if fname == 'result.csv':
        return 'naver_gfa'
    if 'daily-' in fname and fname.endswith('.xlsx'):
        return 'twitter'
    if fname.startswith('dmc_') and '보고서' in fname:
        return 'kakao'

    # 2순위: 인코딩 + 컬럼명
    enc = chardet.detect(file_bytes[:3000])['encoding'] or 'utf-8'
    try:
        if 'utf-16' in enc.lower():
            text = file_bytes.decode('utf-16')
        else:
            text = file_bytes.decode('utf-8-sig', errors='ignore')
        lines = text.split('\n')
        header = lines[2] if len(lines) > 2 else lines[0]

        if 'TrueView 조회수' in header or '동영상 25% 재생' in header:
            return 'youtube'
        if '광고 소재 이름' in header and '총 비용' in header:
            return 'naver_gfa'
        if '소재 이름' in header and '클릭률' in header:
            return 'kakao'
        if '트윗 참여' in header or '리트윗 수' in header:
            return 'twitter'
    except:
        pass

    # xlsx 컬럼 확인
    if fname.endswith('.xlsx'):
        try:
            df = pd.read_excel(io.BytesIO(file_bytes), header=0)
            cols = ' '.join(df.columns.tolist())
            if '트윗 참여' in cols: return 'twitter'
            if 'TrueView' in cols: return 'youtube'
            if '광고 소재 이름' in cols: return 'naver_gfa'
            if '소재 이름' in cols: return 'kakao'
        except:
            pass

    return 'unknown'

MEDIA_INFO = {
    'youtube':   {'label': 'Youtube',    'emoji': '▶', 'color': '#f87171'},
    'naver_gfa': {'label': 'Naver GFA',  'emoji': 'N', 'color': '#4ade80'},
    'kakao':     {'label': 'Kakao',      'emoji': 'K', 'color': '#fbbf24'},
    'twitter':   {'label': 'Twitter (X)','emoji': 'X', 'color': '#4f9cf9'},
    'unknown':   {'label': '알 수 없음',  'emoji': '?', 'color': '#8892a4'},
}

COST_FORMULA = {
    'youtube':   lambda v: f'={v}/0.9',
    'naver_gfa': lambda v: f'={v}/1.1*0.95',
    'kakao':     lambda v: f'={v}*0.95',
    'twitter':   lambda v: f'={v}/0.9',
}

# ─────────────────────────────────────────────
# 파서
# ─────────────────────────────────────────────
CREATIVE_KO = {
    'Snackwrap':'스낵랩','Churros':'츄러스','Churros1':'츄러스1',
    'Churros2':'츄러스2','Shake':'쉐이크','Platform':'플랫폼',
    'Carousel':'DA','DA':'DA',
}
WEEK_RANGES = [
    ('W1','2026-02-03','2026-02-11'),('W2','2026-02-12','2026-02-19'),
    ('W3','2026-02-20','2026-02-25'),('W4','2026-02-26','2026-03-04'),
    ('W5','2026-03-05','2026-03-11'),('W6','2026-03-12','2026-03-18'),
    ('W7','2026-03-19','2026-03-25'),('W8','2026-03-26','2026-04-01'),
    ('W9','2026-04-02','2026-04-08'),('W10','2026-04-09','2026-04-15'),
    ('W11','2026-04-16','2026-04-22'),('W12','2026-04-23','2026-04-29'),
    ('W13','2026-04-30','2026-05-06'),('W14','2026-05-07','2026-05-13'),
    ('W15','2026-05-14','2026-05-20'),
]

def get_week(d):
    d = str(d)[:10].replace('.','-').rstrip('-')
    for w,s,e in WEEK_RANGES:
        if s <= d <= e: return w
    return ''

def parse_date(d):
    return str(d)[:10].replace('.','-').rstrip('-')

def clean_num(v):
    v = str(v).strip().strip('"').replace(',','').replace('%','').strip()
    return '' if v in ['-','nan','--',''] else v

def parse_campaign_code(code, media_map, product_map):
    parts = str(code).split('_')
    if len(parts) < 3: return '', ''
    return media_map.get(parts[1],''), product_map.get(parts[2],'')

def parse_creative_code(code):
    parts = str(code).split('_')
    if len(parts) < 4: return '', ''
    creative = CREATIVE_KO.get(parts[1], parts[1])
    sec = parts[3] if len(parts) > 3 and parts[3].isdigit() else ''
    return creative, sec

RAW_COLS = ['미디어','상품명','캠페인','그룹','광고','소재','초수','일자',
            '비용(net)','노출','클릭','조회','25','50','75','100',
            '0.25','0.5','0.75','1','인게이지','리트윗','답글수','팔로우','미디어참여','투표','Week']
AUTO_COLS = {'미디어','상품명','소재','초수','Week'}

COL_MAP = {
    'youtube':   {'campaign':'캠페인','group':'광고그룹','ad':'광고 이름','date':'일','cost':'net',
                  'imp':'노출수','click':'클릭수','view':'TrueView 조회수',
                  'm25':'동영상 25% 재생','m50':'동영상 50% 재생','m75':'동영상 75% 재생','m100':'동영상 100% 재생'},
    'naver_gfa': {'campaign':'캠페인 이름','group':'광고 그룹 이름','ad':'광고 소재 이름',
                  'date':'기간','cost':'총 비용','imp':'노출','click':'클릭'},
    'kakao':     {'campaign':'캠페인 이름','group':'광고그룹 이름','ad':'소재 이름',
                  'date':'일','cost':'비용','imp':'노출수','click':'클릭수'},
    'twitter':   {'campaign':'캠페인 이름','group':'광고 그룹 이름','ad':'광고 이름',
                  'date':'기간','cost':'비용','imp':'노출수','click':'클릭수',
                  'engage':'트윗 참여','retweet':'리트윗 수','reply':'답글',
                  'follow':'팔로우 수','media_eng':'미디어 참여'},
}

def parse_file(media_key, file_bytes, media_map, product_map):
    rows = []
    cm = COL_MAP.get(media_key, {})

    if media_key == 'youtube':
        enc = chardet.detect(file_bytes[:3000])['encoding'] or 'utf-16'
        text = file_bytes.decode('utf-16' if 'utf-16' in enc.lower() else enc)
        lines = text.split('\n')
        headers = lines[2].strip().split('\t')
        for line in lines[3:]:
            cols = line.strip().split('\t')
            if len(cols) < 10: continue
            row = {headers[i]: cols[i] for i in range(min(len(headers),len(cols)))}
            creative_code = str(row.get('캠페인','')).strip()
            ad_name = str(row.get('광고 이름','')).strip()
            creative, sec = parse_creative_code(creative_code)
            product = next((v for k,v in {'VVC':'VVC','VRC':'VRC','FPM':'FPM','Bumper':'Bumper'}.items() if ad_name.startswith(k)), '')
            date_raw = parse_date(row.get('일',''))
            out = {c:'' for c in RAW_COLS}
            out.update({'미디어':'Youtube','상품명':product,'캠페인':creative_code,
                        '그룹':str(row.get('광고그룹','')).strip(),'광고':ad_name,
                        '소재':creative,'초수':sec,'일자':date_raw,
                        '비용(net)':clean_num(row.get('net','')),
                        '노출':clean_num(row.get('노출수','')),'클릭':clean_num(row.get('클릭수','')),
                        '조회':clean_num(row.get('TrueView 조회수','')),
                        '25':clean_num(row.get('동영상 25% 재생','')),'50':clean_num(row.get('동영상 50% 재생','')),
                        '75':clean_num(row.get('동영상 75% 재생','')),'100':clean_num(row.get('동영상 100% 재생','')),
                        'Week':get_week(date_raw)})
            rows.append(out)

    elif media_key == 'naver_gfa':
        df = pd.read_csv(io.BytesIO(file_bytes), encoding='utf-8-sig')
        df = df[df[cm['campaign']].str.contains('BA Q1', na=False)]
        for _, row in df.iterrows():
            campaign = str(row.get(cm['campaign'],''))
            ad = str(row.get(cm['ad'],''))
            media, product = parse_campaign_code(campaign, media_map, product_map)
            creative, sec = parse_creative_code(ad)
            date_raw = parse_date(row.get(cm['date'],''))
            cost = clean_num(row.get(cm['cost'],''))
            out = {c:'' for c in RAW_COLS}
            out.update({'미디어':media or 'Naver GFA','상품명':product,'캠페인':campaign,
                        '그룹':str(row.get(cm['group'],'')),'광고':ad,
                        '소재':creative,'초수':sec,'일자':date_raw,
                        '비용(net)':cost,'노출':clean_num(row.get(cm['imp'],'')),
                        '클릭':clean_num(row.get(cm['click'],'')), 'Week':get_week(date_raw)})
            rows.append(out)

    elif media_key == 'kakao':
        enc = chardet.detect(file_bytes[:3000])['encoding'] or 'utf-16'
        text = file_bytes.decode('utf-16' if 'utf-16' in enc.lower() else enc)
        df = pd.read_csv(io.StringIO(text), sep='\t', quotechar='"')
        for _, row in df.iterrows():
            campaign = str(row.get(cm['campaign'],''))
            ad = str(row.get(cm['ad'],''))
            media, product = parse_campaign_code(campaign, media_map, product_map)
            creative, sec = parse_creative_code(ad)
            date_raw = parse_date(row.get(cm['date'],''))
            cost = clean_num(str(row.get(cm['cost'],'')))
            out = {c:'' for c in RAW_COLS}
            out.update({'미디어':media or 'Kakao Moment','상품명':product,'캠페인':campaign,
                        '그룹':str(row.get(cm['group'],'')),'광고':ad,
                        '소재':creative,'초수':sec,'일자':date_raw,
                        '비용(net)':cost,'노출':clean_num(str(row.get(cm['imp'],''))),'클릭':clean_num(str(row.get(cm['click'],''))),
                        'Week':get_week(date_raw)})
            rows.append(out)

    elif media_key == 'twitter':
        df = pd.read_excel(io.BytesIO(file_bytes), header=0)
        for _, row in df.iterrows():
            campaign = str(row.get(cm['campaign'],''))
            ad = str(row.get(cm['ad'],''))
            date_raw = parse_date(row.get(cm['date'],''))
            cost = clean_num(str(row.get(cm['cost'],'')))
            group = str(row.get(cm['group'],''))
            product = 'Vertical Video Takeover' if 'VVT' in campaign or 'Vertical' in group else 'PTW'
            out = {c:'' for c in RAW_COLS}
            out.update({'미디어':'Twitter','상품명':product,'캠페인':campaign,
                        '그룹':group,'광고':ad,'소재':'DA','일자':date_raw,
                        '비용(net)':cost,'노출':clean_num(str(row.get(cm['imp'],''))),'클릭':clean_num(str(row.get(cm['click'],''))),
                        '인게이지':clean_num(str(row.get(cm['engage'],''))),
                        '리트윗':clean_num(str(row.get(cm['retweet'],''))),
                        '답글수':clean_num(str(row.get(cm['reply'],''))),
                        '팔로우':clean_num(str(row.get(cm['follow'],''))),
                        '미디어참여':clean_num(str(row.get(cm['media_eng'],''))),
                        'Week':get_week(date_raw)})
            rows.append(out)

    return rows

# ─────────────────────────────────────────────
# 엑셀 출력
# ─────────────────────────────────────────────
def to_excel_bytes(rows, media_key):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Raw Data'
    hf_auto  = PatternFill('solid', start_color='FFF0C040')
    hf_input = PatternFill('solid', start_color='FF1F3D6B')
    hfa  = Font(name='Arial', bold=True, color='FF000000', size=10)
    hfi  = Font(name='Arial', bold=True, color='FFFFFFFF', size=10)
    af   = PatternFill('solid', start_color='FFFFF8E0')
    afont= Font(name='Arial', size=9, color='FF805500')
    nfont= Font(name='Arial', size=9)
    cfont= Font(name='Arial', size=9, color='FF000080')
    thin = Side(style='thin', color='FFCCCCCC')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, cn in enumerate(RAW_COLS, 1):
        cell = ws.cell(row=1, column=ci, value=cn)
        cell.fill = hf_auto if cn in AUTO_COLS else hf_input
        cell.font = hfa if cn in AUTO_COLS else hfi
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bdr

    cost_fn = COST_FORMULA.get(media_key)
    for ri, rd in enumerate(rows, 2):
        for ci, cn in enumerate(RAW_COLS, 1):
            val = rd[cn]
            if cn == '비용(net)' and cost_fn and val:
                try:
                    cell = ws.cell(row=ri, column=ci, value=cost_fn(float(val)))
                    cell.font = cfont
                except:
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.font = nfont
            else:
                if cn in ('노출','클릭','조회') and val:
                    try: val = int(float(val))
                    except: pass
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill = af if cn in AUTO_COLS else PatternFill()
                cell.font = afont if cn in AUTO_COLS else nfont
            cell.border = bdr
            cell.alignment = Alignment(vertical='center')

    col_widths = {'미디어':14,'상품명':16,'캠페인':34,'그룹':24,'광고':22,'소재':10,'초수':6,
                  '일자':12,'비용(net)':18,'노출':12,'클릭':10,'조회':10,'25':8,'50':8,'75':8,
                  '100':8,'0.25':6,'0.5':6,'0.75':6,'1':6,'인게이지':10,'리트윗':8,
                  '답글수':8,'팔로우':8,'미디어참여':10,'투표':8,'Week':8}
    for ci, cn in enumerate(RAW_COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(cn, 10)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ─────────────────────────────────────────────
# session_state 초기화
# ─────────────────────────────────────────────
if 'uploaded_media' not in st.session_state:
    st.session_state.uploaded_media = {}   # {media_key: {rows, filename, time, row_count}}
if 'toasts' not in st.session_state:
    st.session_state.toasts = []           # 최근 업데이트 알림 목록

# ─────────────────────────────────────────────
# 사이드바
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div class="section-label">코드 사전</div>', unsafe_allow_html=True)
    dict_file = st.file_uploader(
        "맥도날드_자동화_작업용.xlsx 업로드",
        type=['xlsx'],
        key='dict_uploader',
        help="Media, Product 코드 사전 파일"
    )
    media_map, product_map = {}, {}
    if dict_file:
        media_map, product_map = load_code_dict(dict_file)
        st.success(f"✓ 코드 사전 로드 완료\n매체 {len(media_map)}개 / 상품 {len(product_map)}개")
    else:
        st.info("코드 사전 파일을 업로드하면\n텍소노미 자동매핑이 활성화됩니다.")

    st.divider()
    st.markdown('<div class="section-label">업로드 현황</div>', unsafe_allow_html=True)

    all_media = ['youtube','naver_gfa','kakao','twitter']
    for mk in all_media:
        info = MEDIA_INFO[mk]
        uploaded = mk in st.session_state.uploaded_media
        if uploaded:
            data = st.session_state.uploaded_media[mk]
            st.markdown(f"""
            <div class="media-card active">
                <div class="media-card-name">{info['emoji']} {info['label']}</div>
                <div class="media-card-status">✓ {data['row_count']}행 업데이트</div>
                <div class="media-card-time">{data['time']}</div>
            </div>""", unsafe_allow_html=True)
        else:
            st.markdown(f"""
            <div class="media-card">
                <div class="media-card-name">{info['emoji']} {info['label']}</div>
                <div class="media-card-status">대기 중</div>
            </div>""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# 메인 영역
# ─────────────────────────────────────────────
st.markdown("""
<div class="app-header">
    <div class="app-badge">McD</div>
    <div>
        <p class="app-title">RAW Data Automator</p>
        <p class="app-sub">매체 리포트 → RAW 데이터 자동 변환</p>
    </div>
</div>
""", unsafe_allow_html=True)

# 업데이트 알림 토스트
if st.session_state.toasts:
    for toast in st.session_state.toasts[-3:]:  # 최근 3개만
        info = MEDIA_INFO.get(toast['media_key'], MEDIA_INFO['unknown'])
        st.markdown(f"""
        <div class="toast">
            <div class="toast-icon">✅</div>
            <div class="toast-text">
                <div class="toast-title">{info['label']} 데이터 업데이트 완료</div>
                <div class="toast-sub">{toast['filename']} · {toast['row_count']}행 · {toast['time']}</div>
            </div>
        </div>""", unsafe_allow_html=True)

# 파일 업로드 영역
st.markdown('<div class="section-label">매체 파일 업로드</div>', unsafe_allow_html=True)
st.caption("파일명 또는 내부 컬럼을 분석하여 매체를 자동으로 감지합니다.")

uploaded_files = st.file_uploader(
    "매체 리포트 파일을 업로드하세요 (여러 파일 동시 가능)",
    type=['csv', 'xlsx', 'xls'],
    accept_multiple_files=True,
    key='media_uploader'
)

if uploaded_files:
    for uf in uploaded_files:
        file_bytes = uf.read()
        media_key = detect_media(uf.name, file_bytes)
        info = MEDIA_INFO[media_key]

        col1, col2 = st.columns([3, 1])
        with col1:
            if media_key == 'unknown':
                st.warning(f"⚠️ `{uf.name}` — 매체를 자동 감지하지 못했습니다. 아래에서 직접 선택해주세요.")
                media_key = st.selectbox(
                    "매체 선택",
                    options=list(MEDIA_INFO.keys())[:-1],
                    format_func=lambda x: MEDIA_INFO[x]['label'],
                    key=f'select_{uf.name}'
                )
            else:
                st.success(f"**{info['label']}** 파일 감지 — `{uf.name}`")

        with col2:
            if st.button("변환 실행", key=f'btn_{uf.name}'):
                with st.spinner(f"{info['label']} 데이터 처리 중..."):
                    try:
                        rows = parse_file(media_key, file_bytes, media_map, product_map)
                        now = datetime.now().strftime('%m/%d %H:%M')
                        st.session_state.uploaded_media[media_key] = {
                            'rows': rows,
                            'filename': uf.name,
                            'time': now,
                            'row_count': len(rows),
                            'media_key': media_key,
                        }
                        st.session_state.toasts.append({
                            'media_key': media_key,
                            'filename': uf.name,
                            'row_count': len(rows),
                            'time': now,
                        })
                        st.rerun()
                    except Exception as e:
                        st.error(f"오류 발생: {e}")

st.divider()

# 결과 및 다운로드 영역
if st.session_state.uploaded_media:
    st.markdown('<div class="section-label">변환 결과</div>', unsafe_allow_html=True)

    tabs = st.tabs([MEDIA_INFO[mk]['label'] for mk in st.session_state.uploaded_media.keys()])
    for tab, (mk, data) in zip(tabs, st.session_state.uploaded_media.items()):
        with tab:
            col1, col2, col3 = st.columns(3)
            col1.metric("총 행수", f"{data['row_count']}행")
            col2.metric("파일명", data['filename'][:20])
            col3.metric("업데이트", data['time'])

            df_preview = pd.DataFrame(data['rows'])
            st.dataframe(
                df_preview[['미디어','상품명','캠페인','소재','초수','일자','Week','비용(net)','노출','클릭']].head(20),
                use_container_width=True,
                height=300
            )

            excel_bytes = to_excel_bytes(data['rows'], mk)
            today = datetime.now().strftime('%Y%m%d')
            st.download_button(
                label=f"⬇  {MEDIA_INFO[mk]['label']} RAW 데이터 다운로드",
                data=excel_bytes,
                file_name=f"RAW_Data_{MEDIA_INFO[mk]['label']}_{today}.xlsx",
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                key=f'download_{mk}'
            )
else:
    st.markdown("""
    <div style="text-align:center;padding:48px;color:#505869;border:1px dashed #252a38;">
        <div style="font-size:32px;margin-bottom:12px;">📂</div>
        <div style="font-family:'IBM Plex Mono',monospace;font-size:12px;">
            매체 파일을 업로드하면 여기에 결과가 표시됩니다
        </div>
    </div>
    """, unsafe_allow_html=True)
